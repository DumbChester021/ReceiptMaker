import os
import openpyxl
import os
import openpyxl

class Product:
    def __init__(self, name, quantity, price):
        self.name = name
        self.quantity = quantity
        self.price = price

class Inventory:
    def __init__(self):
        self.products = {}
        if os.path.exists("inventory.xlsx"):
            workbook = openpyxl.load_workbook('inventory.xlsx')
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                quantity = sheet.cell(row=row, column=2).value
                price = sheet.cell(row=row, column=3).value
                self.products[name] = Product(name, quantity, price)
        else:
            print("Inventory file doesn't exist")

    def add(self, name, quantity, price):
        name = name.lower()
        if os.path.exists("inventory.xlsx"):
            workbook = openpyxl.load_workbook('inventory.xlsx')
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                product_name = sheet.cell(row=row, column=1).value
                if product_name.lower() == name:
                    product_quantity = sheet.cell(row=row, column=2).value
                    sheet.cell(row=row, column=2).value = product_quantity + quantity
                    workbook.save('inventory.xlsx')
                    print("Item Added Succesfully")
                    return
            # If the product does not exist in the inventory
            sheet.append([name, quantity, price])
            workbook.save('inventory.xlsx')
            print("Item Added Succesfully")
        else:
            print("Inventory file doesn't exist")


    def sold(self, name, quantity):
        name = name.lower()
        if os.path.exists("inventory.xlsx"):
            workbook = openpyxl.load_workbook('inventory.xlsx')
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                name_in_file = sheet.cell(row=row, column=1).value.lower()
                if name_in_file == name:
                    quantity_in_file = sheet.cell(row=row, column=2).value
                    if quantity_in_file < quantity:
                        print("Not enough quantity in stock")
                    else:
                        sheet.cell(row=row, column=2).value = quantity_in_file - quantity
                        workbook.save('inventory.xlsx')
                        print("Item Sold Succesfully")
                        return
            print(f"{name} not found in inventory")
        else:
            print("Inventory file doesn't exist")

    def check(self, name):
        name = name.lower()
        if os.path.exists("inventory.xlsx"):
            workbook = openpyxl.load_workbook('inventory.xlsx')
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                name_in_file = sheet.cell(row=row, column=1).value.lower()
                if name_in_file == name:
                    quantity = sheet.cell(row=row, column=2).value
                    price = sheet.cell(row=row, column=3).value
                    print(f"Name: {name_in_file}, Quantity: {quantity}, Price: {price}")
                    return
            print(f"{name} not found in inventory")
        else:
            print("Inventory file doesn't exist")

    def check_all(self):
        self.products = {}
        if os.path.exists("inventory.xlsx"):
            workbook = openpyxl.load_workbook('inventory.xlsx')
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                quantity = sheet.cell(row=row, column=2).value
                price = sheet.cell(row=row, column=3).value
                self.products[name] = Product(name, quantity, price)
        else:
            print("Inventory file doesn't exist")
        if not self.products:
            print("Inventory is Empty")
        else:
            print("Name, Quantity, Price")
            for product in self.products.values():
                print(f"{product.name}, {product.quantity}, {product.price}")

    def export(self, filename):
        filename = filename.lower()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Inventory"
        sheet.cell(row=1, column=1).value = "Product Name"
        sheet.cell(row=1, column=2).value = "Quantity"
        sheet.cell(row=1, column=3).value = "Price"

        row = 2
        for product in self.products.values():
            sheet.cell(row=row, column=1).value = product.name
            sheet.cell(row=row, column=2).value = product.quantity
            sheet.cell(row=row, column=3).value = product.price
            row += 1

        workbook.save(filename)
        print("File Saved Succesfully")

    def soldff(self, filename):
        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                quantity = sheet.cell(row=row, column=2).value
                price = sheet.cell(row=row, column=3).value
                self.sold(name, quantity)
                sales_sheet = workbook.create_sheet(title="Sales")
                sales_sheet.cell(row=row, column=1).value = name
                sales_sheet.cell(row=row, column=2).value = quantity
                sales_sheet.cell(row=row, column=3).value = price
            workbook.save(filename)
            print("Items Sold Succesfully")
        else:
            print(f"{filename} does not exist.")


inv = Inventory()
while True:
    command = input("Enter command (add, sold, check, checkall, soldff, exit): ")
    command = command.lower()
    if command == "add":
        name = input("Enter product name: ")
        quantity = int(input("Enter quantity: "))
        price = float(input("Enter price per item: "))
        inv.add(name, quantity, price)
    elif command == "sold":
        name = input("Enter product name: ")
        quantity = int(input("Enter quantity: "))
        inv.sold(name, quantity)
    elif command == "check":
        name = input("Enter product name: ")
        inv.check(name)
    elif command == "checkall":
        inv.check_all()
    elif command == "exit":
        break
    elif command == "soldff":
        filename = input("Please input Filename: ")
        inv.soldff(filename,filename2)
    else:
        print("Invalid command")
